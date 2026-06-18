"""
excel_writer.py — Clinical Records → Styled Excel Workbook
===========================================================
Writes the standardised output to an Excel file with:
  Sheet 1 — Clinical Records  (one row per test)
  Sheet 2 — Normalization Audit  (rows where confidence < 0.9)
  Sheet 3 — Summary Stats
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import pandas as pd

log = logging.getLogger(__name__)

# ── Output column order (matches the BigQuery schema screenshot) ───────────────
OUTPUT_COLUMNS = [
    "document_id",
    "source_file",
    "patient_name",
    "uhid",
    "age",
    "gender",
    "hospital_name",
    "bill_date",
    "reports_date",
    "admission_date",
    "discharge_date",
    "test_name_original",
    "test_name_canonical",
    "result_value",
    "result_text",
    "unit_original",
    "unit_canonical",
    "range_low",
    "range_high",
    "range_text",
    "test_analytics",
    "normalization_method",
    "normalization_confidence",
]

# ── Column display widths (characters) ────────────────────────────────────────
_COL_WIDTHS = {
    "document_id":              20,
    "source_file":              25,
    "patient_name":             22,
    "uhid":                     18,
    "age":                       6,
    "gender":                    9,
    "hospital_name":            28,
    "bill_date":                13,
    "reports_date":             13,
    "admission_date":           14,
    "discharge_date":           14,
    "test_name_original":       35,
    "test_name_canonical":      28,
    "result_value":             14,
    "result_text":              20,
    "unit_original":            15,
    "unit_canonical":           15,
    "range_low":                11,
    "range_high":               11,
    "range_text":               15,
    "test_analytics":           22,
    "normalization_method":     18,
    "normalization_confidence": 22,
}

# ── Palette ───────────────────────────────────────────────────────────────────
HEADER_FILL  = "1F3864"   # dark navy
HEADER_FONT  = "FFFFFF"   # white
TIER1_FILL   = "D9EAD3"   # light green
TIER2_FILL   = "FFF2CC"   # light amber
TIER3_FILL   = "FCE5CD"   # light orange
UNRES_FILL   = "F4CCCC"   # light red
AUDIT_FILL   = "FFE0B2"   # audit sheet row bg
ALT_ROW_FILL = "F8F9FA"   # alternate row stripe


def _method_fill(method: str) -> str:
    m = (method or "").strip()
    if m.startswith("Tier1"):
        return TIER1_FILL
    if m.startswith("Tier2"):
        return TIER2_FILL
    if m.startswith("Tier3"):
        return TIER3_FILL
    return UNRES_FILL


def write_excel(rows: list[dict], output_path: str | Path) -> Path:
    """
    Write clinical records to a styled .xlsx workbook.
    Returns the resolved output path.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
            GradientFill,
        )
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("openpyxl is required — run: pip install openpyxl") from exc

    output_path = Path(output_path)
    df = pd.DataFrame(rows)

    # Ensure all expected columns exist (fill missing with None)
    for col in OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = None
    df = df[OUTPUT_COLUMNS]

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 1 — Clinical Records
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Clinical Records"
    ws.freeze_panes = "A2"      # freeze header row

    # ── Style helpers ──
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(name="Calibri", bold=True, color=HEADER_FONT, size=10)
    cell_font   = Font(name="Calibri", size=9)
    center_al   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_al     = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    thin_side   = Side(style="thin", color="D0D0D0")
    thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side)

    fill_t1   = PatternFill("solid", fgColor=TIER1_FILL)
    fill_t2   = PatternFill("solid", fgColor=TIER2_FILL)
    fill_t3   = PatternFill("solid", fgColor=TIER3_FILL)
    fill_unr  = PatternFill("solid", fgColor=UNRES_FILL)
    fill_alt  = PatternFill("solid", fgColor=ALT_ROW_FILL)

    def method_fill(method: str):
        m = (method or "").strip()
        if m.startswith("Tier1"):  return fill_t1
        if m.startswith("Tier2"):  return fill_t2
        if m.startswith("Tier3"):  return fill_t3
        return fill_unr

    # ── Header row ──
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center_al
        cell.border = thin_border
        # Column width
        width = _COL_WIDTHS.get(col_name, 16)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20

    # ── Data rows ──
    method_col_idx = OUTPUT_COLUMNS.index("normalization_method") + 1

    for row_idx, record in enumerate(df.itertuples(index=False), start=2):
        method = str(getattr(record, "normalization_method", "") or "")
        row_fill = method_fill(method)
        alt_fill = fill_alt if row_idx % 2 == 0 else None

        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            val = getattr(record, col_name, None)
            # Convert NaN / None to ""
            if val != val or val is None:    # NaN check
                val = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font   = cell_font
            cell.border = thin_border

            # Highlight the normalization_method column by tier colour
            if col_idx == method_col_idx:
                cell.fill = row_fill
                cell.alignment = center_al
            elif col_name in ("result_value", "range_low", "range_high",
                              "normalization_confidence"):
                cell.alignment = center_al
                cell.fill = alt_fill or PatternFill()
            else:
                cell.alignment = left_al
                cell.fill = alt_fill or PatternFill()

        ws.row_dimensions[row_idx].height = 15

    # Add AutoFilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}{len(df)+1}"

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 2 — Normalization Audit (confidence < 0.9 or Unresolved)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Normalization Audit")
    ws2.freeze_panes = "A2"

    audit_cols = [
        "document_id", "source_file", "patient_name",
        "test_name_original", "test_name_canonical",
        "normalization_method", "normalization_confidence",
        "result_value", "unit_canonical",
    ]
    audit_df = df[
        (df["normalization_confidence"].fillna(0).astype(float) < 0.90) |
        (df["normalization_method"] == "Unresolved")
    ][audit_cols].copy()

    audit_header_fill = PatternFill("solid", fgColor="7B2D8B")
    audit_header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    audit_row_fill    = PatternFill("solid", fgColor=AUDIT_FILL)

    audit_widths = {
        "document_id": 20, "source_file": 22, "patient_name": 20,
        "test_name_original": 35, "test_name_canonical": 28,
        "normalization_method": 18, "normalization_confidence": 22,
        "result_value": 14, "unit_canonical": 15,
    }
    for col_idx, col_name in enumerate(audit_cols, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.fill = audit_header_fill
        cell.font = audit_header_font
        cell.alignment = center_al
        ws2.column_dimensions[get_column_letter(col_idx)].width = audit_widths.get(col_name, 18)

    for row_idx, rec in enumerate(audit_df.itertuples(index=False), start=2):
        for col_idx, col_name in enumerate(audit_cols, start=1):
            val = getattr(rec, col_name, None)
            if val != val or val is None:
                val = ""
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font   = cell_font
            cell.fill   = audit_row_fill if row_idx % 2 == 0 else PatternFill()
            cell.border = thin_border
            cell.alignment = center_al if col_name in ("normalization_confidence", "result_value") else left_al

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(audit_cols))}{len(audit_df)+1}"

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 3 — Summary Statistics
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 18

    summary_header_fill = PatternFill("solid", fgColor="2E4057")
    summary_font        = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    value_font          = Font(name="Calibri", size=10)
    section_font        = Font(name="Calibri", bold=True, size=10, color="1F3864")

    def _ws3_title(row, text):
        c = ws3.cell(row=row, column=1, value=text)
        c.font = summary_font
        c.fill = summary_header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws3.merge_cells(f"A{row}:B{row}")
        ws3.row_dimensions[row].height = 20

    def _ws3_row(row, label, value, bold_label=False):
        c1 = ws3.cell(row=row, column=1, value=label)
        c2 = ws3.cell(row=row, column=2, value=value)
        c1.font = Font(name="Calibri", bold=bold_label, size=10)
        c2.font = value_font
        c1.alignment = left_al
        c2.alignment = center_al

    r = 1
    _ws3_title(r, "📊  Schema Engine — Run Summary"); r += 1
    ws3.cell(row=r, column=1, value="").font = value_font; r += 1

    _ws3_title(r, "File Statistics"); r += 1
    total_files = df["source_file"].nunique()
    total_rows  = len(df)
    _ws3_row(r, "Total JSON files processed", total_files); r += 1
    _ws3_row(r, "Total test records extracted", total_rows); r += 1
    _ws3_row(r, "Unique canonical test names", df["test_name_canonical"].nunique()); r += 1
    _ws3_row(r, "Unique patients", df["patient_name"].nunique()); r += 1
    r += 1

    _ws3_title(r, "Normalization Method Breakdown"); r += 1
    method_counts = df["normalization_method"].value_counts()
    for method, count in method_counts.items():
        pct = f"{count / total_rows * 100:.1f}%" if total_rows else "0%"
        _ws3_row(r, f"  {method}", f"{count}  ({pct})"); r += 1
    r += 1

    _ws3_title(r, "Confidence Distribution"); r += 1
    conf = df["normalization_confidence"].dropna().astype(float)
    if not conf.empty:
        _ws3_row(r, "  Mean confidence",   f"{conf.mean():.4f}"); r += 1
        _ws3_row(r, "  Median confidence", f"{conf.median():.4f}"); r += 1
        _ws3_row(r, "  Min confidence",    f"{conf.min():.4f}"); r += 1
        hi = (conf >= 0.90).sum()
        lo = (conf < 0.90).sum()
        _ws3_row(r, "  High confidence (≥ 0.90)", f"{hi}  ({hi/total_rows*100:.1f}%)"); r += 1
        _ws3_row(r, "  Needs review (< 0.90)",    f"{lo}  ({lo/total_rows*100:.1f}%)"); r += 1
    r += 1

    _ws3_title(r, "Top 10 Canonical Test Names (by frequency)"); r += 1
    for name, count in df["test_name_canonical"].value_counts().head(10).items():
        _ws3_row(r, f"  {name}", count); r += 1

    # ── Legend ──
    r += 1
    _ws3_title(r, "Legend — Row Colours in Clinical Records Sheet"); r += 1
    for colour, meaning in [
        (TIER1_FILL, "Tier 1 — Canonical Registry Exact Match"),
        (TIER2_FILL, "Tier 2 — NLP Jaccard Similarity Match"),
        (TIER3_FILL, "Tier 3 — Gemini Semantic Fallback"),
        (UNRES_FILL, "Unresolved — Raw name kept as-is"),
    ]:
        c1 = ws3.cell(row=r, column=1, value=f"  {meaning}")
        c1.fill = PatternFill("solid", fgColor=colour)
        c1.font = Font(name="Calibri", size=10)
        c1.alignment = left_al
        ws3.merge_cells(f"A{r}:B{r}")
        r += 1

    wb.save(output_path)
    log.info("Excel written → %s  (%d rows, %d audit rows)",
             output_path, len(df), len(audit_df))
    return output_path
